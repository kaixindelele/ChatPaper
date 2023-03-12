# -*- coding: utf-8 -*-
import pandas as pd
import os
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

def OutputExcel(df,pdfdata,filename):
    df = pd.DataFrame.from_dict(df, orient='index')
    # 将行和列翻转
    df = df.T
    sheet_name = 'Sheet'
    # 判断Excel文件是否存在，不存在则创建新文件
    if os.path.isfile(filename):
        book = load_workbook(filename)
        # 判断工作表是否存在，不存在则创建新工作表
        if sheet_name in book.sheetnames:
            startrow = book[sheet_name].max_row  #获取目前最大行数
            df['Serial Number'] = startrow
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            writer.book = book
            writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
            startrow = writer.sheets[sheet_name].max_row
            # 将数据追加到现有文件
            df.to_excel(writer, sheet_name=sheet_name, index=False,index_label='Key', startrow=startrow, header=False)
        else:
            df['Serial Number'] = 1
            writer = pd.ExcelWriter(filename, engine='openpyxl')
            writer.book = book
            df.to_excel(writer, sheet_name=sheet_name, index=False, index_label='Key', startrow=0, header=True)
    else:
        df['Serial Number'] = 1
        book = Workbook()
        # 删除默认创建的工作表
        default_sheet = book['Sheet']
        book.remove(default_sheet)
        # 创建新工作表
        writer = pd.ExcelWriter(filename, engine='openpyxl')
        writer.book = book
        df.to_excel(writer, sheet_name=sheet_name, index=False, index_label='Key', startrow=0, header=True)

    # 设置Excel单元格格式
    worksheet = writer.sheets[sheet_name]
    for col in worksheet.columns:
        max_length = 0
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        for cell in col:
            cell.alignment = Alignment(wrap_text=True, vertical='center')
        cell_value = col[1].value
        if cell_value is not None and len(str(cell_value)) <= 20:
            adjusted_width = len(str(cell_value)) + 5
        elif cell_value is not None and len(str(cell_value)) <= 200:
            adjusted_width = 20
        else:
            adjusted_width = 50
        worksheet.column_dimensions[col[0].column_letter].width = adjusted_width

    # 添加pdf超链接地址
    startrow = book[sheet_name].max_row  # 获取目前最大行数
    #获取当前文件PDF地址
    with open(pdfdata, 'r', encoding="utf-8") as file:
        pdflocal = file.read()
    ws=book.active
    url = "file:///"+pdflocal
    ws.cell(row=startrow, column=3).hyperlink = url
    # 为单元格添加字体，使得超链接更加明显
    ws.cell(row=startrow, column=3).font = Font(underline='single', color='0563C1')

    # 保存Excel文件
    writer.save()
